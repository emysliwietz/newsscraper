import os
import string
import xlrd
import openpyxl

files_relaxation = [
    'corona_lockerung_de',
    'corona_relaxation_en',
    'corona_versoepeling_nl'
]

files_measures = [
    'corona_maßnahmen_de',
    'corona_measures_en',
    'corona_maatregelen_nl'
]

files_tracking_app = [
    'corona_tracking_app_de',
    'corona_tracking_app_en',
    'corona_tracking_app_nl'
]

files_mask_requirement = [
    'corona_maskenpflicht_de',
    'corona_mask_requirement_en',
    'corona_masker_nl'
]

files_donald_trump = [
    'donald_trump_de',
    'donald_trump_en',
    'donald_trump_nl'
]

files_eu = [
    'eu_de',
    'eu_en',
    'eu_nl'
]

files_italy = [
    'italien_de',
    'italy_en',
    'italie_nl'
]

files_who = [
    'who_de',
    'who_en_with_lines_removed',
    'who_nl'
]

#alphabet = list(string.ascii_uppercase)

if not os.path.exists("analysis"):
    os.mkdir("analysis")

if not os.path.exists("analysis/atlasti_codes"):
    os.mkdir("analysis/atlasti_codes")


def code_scan():
    workbook = xlrd.open_workbook('Bachelor_thesis_CD.xlsx')
    worksheet = workbook.sheet_by_name('CodeDocumentTable')
    for column in range(1, 25):
        file = worksheet.cell(0, column).value.split('\n')[0]
        print(file)
        if file == "who_en_with_lines_removed":
            language = "EN"
        else:
            language = file.split('_')[len(file.split('_')) - 1].upper()
        print(language)
        topic = ""
        string_to_append = ""
        if file in files_relaxation:
            topic = "relaxation"
        elif file in files_measures:
            topic = "measures"
        elif file in files_tracking_app:
            topic = "tracking_app"
        elif file in files_mask_requirement:
            topic = "mask_requirement"
        elif file in files_donald_trump:
            topic = "donald_trump"
        elif file in files_eu:
            topic = "eu"
        elif file in files_italy:
            topic = "italy"
        elif file in files_who:
            topic = "who"

        filepath = os.path.join("analysis/atlasti_codes", topic + "_codes.txt")
        line = "{}: \n".format(language)
        string_to_append += line
        write_mode = "a"  # Append to file
        char = ""
        codes_to_append = ""
        for row in range(1, 297):
            code = worksheet.cell(row, 0).value.split("\n")[0]
            char = code[0:2]
            if int(worksheet.cell(row, column).value) > 0:
                codes_to_append += code

        code_list = codes_to_append.split(char)
        for code in code_list:
            line = "{}\n".format(str(code))
            string_to_append += line
        string_to_append += "\n"
        with open(filepath, write_mode, encoding='utf-8') as f:
            f.write(string_to_append)

        print("Finished checking codes for " + file)


code_scan()
